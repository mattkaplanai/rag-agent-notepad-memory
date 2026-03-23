"""
Two-level decision cache + Excel export.
Level 1: exact hash match. Level 2: semantic embedding similarity.
"""

import json
import logging
import time
from datetime import datetime
from pathlib import Path

from app.config import CACHE_FILE, EMBEDDING_MODEL, EMBEDDING_TIMEOUT, EXCEL_FILE, SEMANTIC_THRESHOLD
from app.utils import cosine_similarity, hash_inputs

logger = logging.getLogger(__name__)


def _get_embedding(text):
    from openai import OpenAI
    client = OpenAI(timeout=EMBEDDING_TIMEOUT)
    response = client.embeddings.create(input=text, model=EMBEDDING_MODEL)
    return response.data[0].embedding


class DecisionCache:
    def __init__(self, cache_path: Path = CACHE_FILE):
        self.cache_path = cache_path
        self.entries: list[dict] = []
        self._load()

    def _load(self):
        if self.cache_path.exists():
            try:
                data = json.loads(self.cache_path.read_text(encoding="utf-8"))
                self.entries = data if isinstance(data, list) else []
                logger.info("Loaded %d cached decisions.", len(self.entries))
            except Exception:
                self.entries = []

    def _save(self):
        self.cache_path.write_text(json.dumps(self.entries, ensure_ascii=False, indent=2), encoding="utf-8")
        self._export_excel()

    def _export_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Refund Decisions"

        headers = ["No", "Date/Time", "Case Type", "Flight Type", "Ticket Type", "Payment Method",
                    "Description", "Decision", "Confidence", "Analysis Steps", "Reasons",
                    "Applicable Regulations", "Refund Type", "Refund Payment", "Refund Timeline", "Passenger Action Items"]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        decision_fills = {
            "APPROVED": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "DENIED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "PARTIAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        }

        for row_idx, entry in enumerate(self.entries, 2):
            result = entry.get("result", {})
            refund = result.get("refund_details") or {}
            ts = entry.get("timestamp", 0)
            dt_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else ""

            row_data = [
                row_idx - 1, dt_str, entry.get("case_type", ""), entry.get("flight_type", ""),
                entry.get("ticket_type", ""), entry.get("payment_method", ""),
                entry.get("description_preview", ""), result.get("decision", ""), result.get("confidence", ""),
                "\n".join(result.get("analysis_steps", [])), "\n".join(result.get("reasons", [])),
                "\n".join(result.get("applicable_regulations", [])),
                refund.get("refund_type", "N/A"), refund.get("payment_method", "N/A"),
                refund.get("timeline", "N/A"), "\n".join(result.get("passenger_action_items", [])),
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

            fill = decision_fills.get(result.get("decision", ""))
            if fill:
                ws.cell(row=row_idx, column=8).fill = fill
                ws.cell(row=row_idx, column=8).font = Font(bold=True)

        col_widths = [5, 18, 22, 18, 15, 15, 40, 12, 12, 50, 40, 40, 30, 25, 25, 40]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(str(EXCEL_FILE))

    def lookup(self, case_type, flight_type, ticket_type, payment_method, accepted_alternative, description, tenant_id=''):
        """Look up a decision. Returns (result, status, query_embedding).

        The query_embedding is returned so callers can pass it to store()
        without re-computing it (avoids a redundant OpenAI API call).

        Cache keys are scoped per tenant — Delta and United never share cached results.
        """
        input_hash = hash_inputs(case_type, flight_type, ticket_type, payment_method, accepted_alternative, description)
        if tenant_id:
            input_hash = f"{tenant_id}:{input_hash}"

        for entry in self.entries:
            if entry.get("hash") == input_hash:
                logger.info("Cache exact hit.")
                return entry["result"], "exact_hit", None

        if not description.strip():
            return None, "miss", None

        query_embedding = _get_embedding(description)
        best_sim, best_entry = 0.0, None
        for entry in self.entries:
            emb = entry.get("embedding")
            if not emb:
                continue
            sim = cosine_similarity(query_embedding, emb)
            if sim > best_sim:
                best_sim = sim
                best_entry = entry

        if best_entry and best_sim >= SEMANTIC_THRESHOLD:
            logger.info("Cache semantic hit (similarity %.3f).", best_sim)
            return best_entry["result"], "semantic_hit", query_embedding

        logger.info("Cache miss (best similarity %.3f).", best_sim)
        return None, "miss", query_embedding

    def store(self, case_type, flight_type, ticket_type, payment_method, accepted_alternative, description, result, embedding=None, tenant_id=''):
        """Store a decision. Pass embedding from lookup() to avoid redundant API call."""
        input_hash = hash_inputs(case_type, flight_type, ticket_type, payment_method, accepted_alternative, description)
        if tenant_id:
            input_hash = f"{tenant_id}:{input_hash}"
        if embedding is None and description.strip():
            embedding = _get_embedding(description)

        self.entries.append({
            "hash": input_hash, "tenant_id": tenant_id,
            "case_type": case_type, "flight_type": flight_type,
            "ticket_type": ticket_type, "payment_method": payment_method,
            "accepted_alternative": accepted_alternative, "description_preview": description[:200],
            "embedding": embedding or [], "result": result, "timestamp": time.time(),
        })
        self._save()
        logger.info("Stored decision (cache now has %d entries).", len(self.entries))

    @property
    def stats(self):
        return {"total_entries": len(self.entries), "cache_file": str(self.cache_path)}

    def clear(self):
        self.entries = []
        self._save()
        if EXCEL_FILE.exists():
            EXCEL_FILE.unlink()
        logger.info("Cache cleared.")
