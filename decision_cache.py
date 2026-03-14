"""
Decision Cache: Two-level caching to minimize LLM costs.

Level 1 — Exact match:  Hash all form inputs. If identical case was seen
                         before, return the cached decision instantly ($0).

Level 2 — Semantic match: Embed the description text and compare cosine
                          similarity to cached descriptions. If similarity
                          exceeds the threshold, reuse the cached decision
                          (costs only 1 embedding call instead of full LLM).

Cache is persisted to disk as JSON (for fast lookups with embeddings)
and also exported to an Excel spreadsheet (for human review).
"""

import hashlib
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parent
CACHE_FILE = PROJECT_ROOT / "decision_cache.json"
EXCEL_FILE = PROJECT_ROOT / "decision_log.xlsx"

SEMANTIC_THRESHOLD = 0.90


def _cosine_similarity(a: list[float], b: list[float]) -> float:
    a_arr = np.array(a)
    b_arr = np.array(b)
    dot = np.dot(a_arr, b_arr)
    norm = np.linalg.norm(a_arr) * np.linalg.norm(b_arr)
    if norm == 0:
        return 0.0
    return float(dot / norm)


def _hash_inputs(
    case_type: str,
    flight_type: str,
    ticket_type: str,
    payment_method: str,
    accepted_alternative: str,
    description: str,
) -> str:
    raw = "|".join([
        case_type.strip().lower(),
        flight_type.strip().lower(),
        ticket_type.strip().lower(),
        payment_method.strip().lower(),
        accepted_alternative.strip().lower(),
        description.strip().lower(),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _get_embedding(text: str) -> list[float]:
    from openai import OpenAI
    client = OpenAI()
    model = os.getenv("OPENAI_EMBEDDING_MODEL", "text-embedding-3-small")
    response = client.embeddings.create(input=text, model=model)
    return response.data[0].embedding


class DecisionCache:
    """Two-level cache: exact hash match → semantic similarity → LLM."""

    def __init__(self, cache_path: Path = CACHE_FILE):
        self.cache_path = cache_path
        self.entries: list[dict] = []
        self._load()

    def _load(self):
        if self.cache_path.exists():
            try:
                data = json.loads(self.cache_path.read_text(encoding="utf-8"))
                self.entries = data if isinstance(data, list) else []
                print(f"[CACHE] Loaded {len(self.entries)} cached decisions.")
            except Exception:
                self.entries = []
        else:
            self.entries = []

    def _save(self):
        n = len(self.entries)
        print(f"[LOG] Cache save: writing JSON to {self.cache_path.name} ({n} entries).", flush=True)
        self.cache_path.write_text(
            json.dumps(self.entries, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        self._export_excel()

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            # Excel export optional; JSON cache still works. Install openpyxl for Excel log.
            return

        n = len(self.entries)
        print(f"[LOG] Excel export: writing {EXCEL_FILE.name} ({n} rows).", flush=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Refund Decisions"

        headers = [
            "No", "Date/Time", "Case Type", "Flight Type", "Ticket Type",
            "Payment Method", "Accepted Alternative", "Description", "Decision", "Confidence",
            "Analysis Steps", "Reasons", "Applicable Regulations",
            "Refund Type", "Refund Payment", "Refund Timeline",
            "Passenger Action Items",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        decision_fills = {
            "APPROVED": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "DENIED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "PARTIAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        }

        for row_idx, entry in enumerate(self.entries, 2):
            result = entry.get("result", {})
            refund = result.get("refund_details") or {}
            ts = entry.get("timestamp", 0)
            dt_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else ""

            row_data = [
                row_idx - 1,
                dt_str,
                entry.get("case_type", ""),
                entry.get("flight_type", ""),
                entry.get("ticket_type", ""),
                entry.get("payment_method", ""),
                entry.get("accepted_alternative", ""),
                entry.get("description_preview", ""),
                result.get("decision", ""),
                result.get("confidence", ""),
                "\n".join(result.get("analysis_steps", [])),
                "\n".join(result.get("reasons", [])),
                "\n".join(result.get("applicable_regulations", [])),
                refund.get("refund_type", "N/A"),
                refund.get("payment_method", "N/A"),
                refund.get("timeline", "N/A"),
                "\n".join(result.get("passenger_action_items", [])),
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

            decision_cell = ws.cell(row=row_idx, column=9)
            fill = decision_fills.get(result.get("decision", ""))
            if fill:
                decision_cell.fill = fill
                decision_cell.font = Font(bold=True)

        col_widths = [5, 18, 22, 18, 15, 15, 28, 40, 12, 12, 50, 40, 40, 30, 25, 25, 40]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(str(EXCEL_FILE))
        print(f"[CACHE] 📊 Excel log updated: {EXCEL_FILE.name} ({len(self.entries)} rows)", flush=True)

    def lookup(
        self,
        case_type: str,
        flight_type: str,
        ticket_type: str,
        payment_method: str,
        accepted_alternative: str,
        description: str,
    ) -> tuple[Optional[dict], str]:
        """
        Look up a cached decision.

        Returns:
            (result_dict, cache_status)
            cache_status is one of: "exact_hit", "semantic_hit", "miss"
        """
        input_hash = _hash_inputs(
            case_type, flight_type, ticket_type,
            payment_method, accepted_alternative, description,
        )
        short_hash = input_hash[:12] if input_hash else "?"
        print("[LOG] Cache: checking exact match (by input hash)...", flush=True)
        # Level 1: exact match
        for entry in self.entries:
            if entry.get("hash") == input_hash:
                print(f"[CACHE] ✅ Exact hit — returning cached decision (saved full LLM call)", flush=True)
                return entry["result"], "exact_hit"
        print("[LOG] Cache exact: miss. Checking cache semantic (embedding + similarity)...", flush=True)

        # Level 2: semantic similarity on description
        if not description.strip():
            print("[LOG] Cache semantic: skipped (no description) — miss.", flush=True)
            return None, "miss"
        query_embedding = _get_embedding(description)

        best_sim = 0.0
        best_entry = None
        for entry in self.entries:
            emb = entry.get("embedding")
            if not emb:
                continue
            if len(emb) != len(query_embedding):
                continue  # different embedding model (e.g. small vs large)
            sim = _cosine_similarity(query_embedding, emb)
            if sim > best_sim:
                best_sim = sim
                best_entry = entry

        if best_entry and best_sim >= SEMANTIC_THRESHOLD:
            print(f"[CACHE] 🔍 Semantic hit — similarity {best_sim:.3f} ≥ {SEMANTIC_THRESHOLD} (saved LLM call)", flush=True)
            return best_entry["result"], "semantic_hit"
        print(f"[LOG] Cache semantic: miss (best_sim={best_sim:.3f} < {SEMANTIC_THRESHOLD}).", flush=True)
        return None, "miss"

    def store(
        self,
        case_type: str,
        flight_type: str,
        ticket_type: str,
        payment_method: str,
        accepted_alternative: str,
        description: str,
        result: dict,
    ):
        """Store a new decision in the cache."""
        input_hash = _hash_inputs(
            case_type, flight_type, ticket_type,
            payment_method, accepted_alternative, description,
        )
        short_hash = input_hash[:12] if input_hash else "?"
        print(f"[LOG] Cache store: adding entry hash={short_hash}... (JSON + Excel).", flush=True)
        embedding = _get_embedding(description) if description.strip() else []

        entry = {
            "hash": input_hash,
            "case_type": case_type,
            "flight_type": flight_type,
            "ticket_type": ticket_type,
            "payment_method": payment_method,
            "accepted_alternative": accepted_alternative,
            "description_preview": description[:200],
            "embedding": embedding,
            "result": result,
            "timestamp": time.time(),
        }
        self.entries.append(entry)
        self._save()
        print(f"[CACHE] 💾 Stored decision — cache now has {len(self.entries)} entries.", flush=True)

    @property
    def stats(self) -> dict:
        return {
            "total_entries": len(self.entries),
            "cache_file": str(self.cache_path),
        }

    def import_from_excel(
        self,
        excel_path: Optional[Path] = None,
        *,
        compute_embeddings: bool = True,
    ) -> tuple[int, int]:
        """
        Seed the cache from an Excel log (e.g. decision_log.xlsx).
        Rows already in the cache (same hash) are skipped.
        Returns (imported_count, skipped_count).
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[CACHE] Excel import requires openpyxl. Run: pip install openpyxl")
            return 0, 0

        path = excel_path or EXCEL_FILE
        if not path.exists():
            print(f"[CACHE] Import skipped: {path} not found.")
            return 0, 0

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        if len(rows) < 2:
            print("[CACHE] Import skipped: Excel has no data rows.")
            return 0, 0

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col = {h: i for i, h in enumerate(header)}

        def get(row: tuple, key: str, default: str = "") -> str:
            i = col.get(key, -1)
            if i < 0 or i >= len(row):
                return default
            v = row[i]
            return str(v).strip() if v is not None else default

        def to_list(s: str) -> list:
            if not s or not s.strip():
                return []
            return [x.strip() for x in s.split("\n") if x.strip()]

        existing_hashes = {e.get("hash") for e in self.entries}
        imported = 0
        skipped = 0

        for row in rows[1:]:
            if not row:
                continue
            case_type = get(row, "Case Type")
            flight_type = get(row, "Flight Type")
            ticket_type = get(row, "Ticket Type")
            payment_method = get(row, "Payment Method")
            accepted_alternative = get(row, "Accepted Alternative")
            description = get(row, "Description")
            if not description and not case_type:
                continue
            input_hash = _hash_inputs(
                case_type, flight_type, ticket_type,
                payment_method, accepted_alternative, description,
            )
            if input_hash in existing_hashes:
                skipped += 1
                continue

            analysis_steps = to_list(get(row, "Analysis Steps"))
            reasons = to_list(get(row, "Reasons"))
            applicable_regulations = to_list(get(row, "Applicable Regulations"))
            passenger_action_items = to_list(get(row, "Passenger Action Items"))
            refund_type = get(row, "Refund Type") or "N/A"
            refund_payment = get(row, "Refund Payment") or "N/A"
            refund_timeline = get(row, "Refund Timeline") or "N/A"
            if refund_type == "N/A" and refund_payment == "N/A" and refund_timeline == "N/A":
                refund_details = {}
            else:
                refund_details = {
                    "refund_type": refund_type,
                    "payment_method": refund_payment,
                    "timeline": refund_timeline,
                }

            result = {
                "decision": get(row, "Decision"),
                "confidence": get(row, "Confidence"),
                "analysis_steps": analysis_steps,
                "reasons": reasons,
                "applicable_regulations": applicable_regulations,
                "refund_details": refund_details,
                "passenger_action_items": passenger_action_items,
            }
            if not result.get("decision"):
                continue

            embedding = []
            if compute_embeddings and description.strip():
                try:
                    embedding = _get_embedding(description)
                except Exception:
                    pass
            entry = {
                "hash": input_hash,
                "case_type": case_type,
                "flight_type": flight_type,
                "ticket_type": ticket_type,
                "payment_method": payment_method,
                "accepted_alternative": accepted_alternative,
                "description_preview": description[:200],
                "embedding": embedding,
                "result": result,
                "timestamp": time.time(),
            }
            self.entries.append(entry)
            existing_hashes.add(input_hash)
            imported += 1

        if imported > 0:
            self._save()
            print(f"[CACHE] 📥 Imported {imported} rows from Excel (skipped {skipped} duplicates).")
        return imported, skipped

    def clear(self):
        self.entries = []
        self._save()
        if EXCEL_FILE.exists():
            EXCEL_FILE.unlink()
        print("[CACHE] 🗑️ Cache and Excel log cleared.")
